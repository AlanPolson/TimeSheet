import os
from openpyxl import load_workbook
import datetime
import re
from collections import OrderedDict 

#identify day of the week, subtract it from current day to get start of current week:
start_date = datetime.date.today() - datetime.timedelta(datetime.date.today().isocalendar()[2]-1)
end_date = start_date + datetime.timedelta(4)
new_filename = start_date.strftime("%Y_%m_%d")+'-'+end_date.strftime("%m_%d")+"_Alan_Time.xlsm"

i=-2
old_worksheet_names=OrderedDict()
ss=load_workbook(r"C:\Users\ampolso\OneDrive - Emory University\2019_01_01-01_01_Alan_Time_Template.xlsm", read_only=False, keep_vba=True)

for worksheet in ss.worksheets:
    i+=1
    if (i>=0 and i<=5):
        #print (i,worksheet,(start_date+datetime.timedelta(i)).strftime("%m-%d %a"))
        old_worksheet_names[worksheet.title]=(start_date+datetime.timedelta(i)).strftime("%m-%d %a")
        for row in worksheet.iter_rows(min_row=0):
            for cell in row:
                if cell.value==None:
                    continue;
                #print ("OLD",cell.value)
                for old_date in old_worksheet_names:
                    #print("Checking",old_date,old_worksheet_names[old_date])
                    cell.value=re.sub(old_date,old_worksheet_names[old_date],cell.value)
                    #print ("NEW",cell.value)
        if (i<5):
            worksheet.title = (start_date+datetime.timedelta(i)).strftime("%m-%d %a")
    
        #print (i,worksheet,old_worksheet_names)            
    ss.save(new_filename)